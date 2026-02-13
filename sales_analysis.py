import pandas as pd

df = pd.read_excel("sales_voucher.xlsx")

#GENERAL ANALYSIS REQUIRED OF MY DATA
print(df.head()) 
print("Shape of data (rows, columns):", df.shape)

print("\nColumn Names:")
print(df.columns)

print("\nData Types:")
print(df.dtypes)

print("\nFirst 5 Rows:")
print(df.head())

print("\nLast 5 Rows:")
print(df.tail())

df["Operator Number"] = df["Operator Number"].astype(str)
df["tax amount"] = df["tax amount"].fillna(0)
df["discount"] = df["discount"].fillna(0)
df["State"] = df["State"].fillna("Unknown")
df["company name"] = df["company name"].fillna("Unknown")

print("\nMissing values in each column:")
print(df.isnull().sum())

print("\nBUSINESS SUMMARY\n")

#TOTAL REVENUE 
total_revenue = df["Recieved amount"].sum()
print(f"Total Revenue: ₹{total_revenue:,.2f}")

#REVENUE BY DATE
daily_sales = df.groupby("Date")["Recieved amount"].sum().sort_index()
print("\nRevenue by Date:")
print(daily_sales)

#REVENUE BY STATE
state_sales = df.groupby("State")["Recieved amount"].sum().sort_values(ascending=False)
print("\nRevenue by State:")
print(state_sales)

#TOP OPERATORS
operator_sales = df.groupby("Operator Name")["Recieved amount"].sum().sort_values(ascending=False)
print("\nTop Operators by Revenue:")
print(operator_sales.head(5))

print("\nRevenue by State:")
state_revenue = df.groupby("State")["Recieved amount"].sum().sort_values(ascending=False)
print(state_revenue)

#TOP 10 TRANSACTIONS
print("\nRevenue per Transaction ID:")
txn_revenue = df.groupby("Txn ID")["Recieved amount"].sum().sort_values(ascending=False)
print(txn_revenue.head(10))  

#DAILY REVENUE
print("\nDaily Revenue Trend:")
daily_revenue = df.groupby("Date")["Recieved amount"].sum().sort_index()
print(daily_revenue)

#BUSINESS SUMMARY CALCULATIONS 
top_operators = df.groupby("Operator Name")["Recieved amount"].sum().sort_values(ascending=False)
state_revenue = df.groupby("State")["Recieved amount"].sum().sort_values(ascending=False)
txn_revenue = df.groupby("Txn ID")["Recieved amount"].sum().sort_values(ascending=False)
daily_revenue = df.groupby("Date")["Recieved amount"].sum().sort_index()

#EXPORT TO MULTI-SHEET EXCEL
with pd.ExcelWriter("sales_report.xlsx", engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Cleaned Data", index=False)
    top_operators.to_excel(writer, sheet_name="Top Operators")
    state_revenue.to_excel(writer, sheet_name="Revenue by State")
    txn_revenue.to_excel(writer, sheet_name="Txn Revenue")
    daily_revenue.to_excel(writer, sheet_name="Daily Trend")

print("Gaurav your multi sheet excel report created")

# CHARTS (For Visulaization)
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, BarChart

wb = load_workbook("sales_report.xlsx")

#LINE CHART FOR DAILY TREND ANALYSIS
ws_trend = wb["Daily Trend"]
chart1 = LineChart()
chart1.title = "Daily Revenue Trend"

data = Reference(ws_trend, min_col=2, min_row=1, max_row=ws_trend.max_row)
cats = Reference(ws_trend, min_col=1, min_row=2, max_row=ws_trend.max_row)

chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
ws_trend.add_chart(chart1, "E2")

#BAR CHART FOR TOP OPERATORS ANALYSIS
ws_ops = wb["Top Operators"]
chart2 = BarChart()
chart2.title = "Top Operators by Revenue"

data2 = Reference(ws_ops, min_col=2, min_row=1, max_row=ws_ops.max_row)
cats2 = Reference(ws_ops, min_col=1, min_row=2, max_row=ws_ops.max_row)

chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
ws_ops.add_chart(chart2, "E2")

wb.save("sales_report.xlsx")

print("Gaurav your charts are added successfully")
